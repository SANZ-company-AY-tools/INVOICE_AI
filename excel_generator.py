"""
Excel and CSV generator module for invoice data export.
Creates clean Excel files and SAP-compatible CSV from extracted invoice data.
Two formats: CON PEDIDO (MIRO) and SIN PEDIDO (FI).
"""

import csv
from datetime import date, datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


# Mapping receptor name keywords → SAP company code (Sociedad)
# Matches ANY name containing the keyword (e.g. "hondakin" matches "Ambar Hondakin SL")
SOCIEDAD_MAP = {
    'eco': 'AE0',
    'plus': 'AP0',
    'hondakin': 'AH0',
    'hodnakin': 'AH0',
    'remasur': 'AR0',
    'pepe': 'AB0',
    'nuñez': 'AB0',
    'nunez': 'AB0',
}


def _get_sociedad(receiver_name: str) -> str:
    """Match receiver name to SAP company code."""
    if not receiver_name:
        return ''
    name_lower = receiver_name.lower().strip()
    for key, code in SOCIEDAD_MAP.items():
        if key in name_lower:
            return code
    return ''


def _format_date_ddmmyyyy(date_str: str) -> str:
    """Convert date string to DD/MM/YYYY format."""
    if not date_str:
        return ''
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime('%d/%m/%Y')
        except (ValueError, AttributeError):
            continue
    return date_str  # Return as-is if no format matches


def _expand_invoices(data: List[Dict]) -> List[Dict]:
    """Expand invoices so each combination of order number + tax line gets its own row."""
    expanded = []
    today = date.today().strftime('%d/%m/%Y')

    for invoice in data:
        if invoice.get('status') != 'success':
            continue

        order_nums = invoice.get('order_numbers', [])
        if not isinstance(order_nums, list):
            order_nums = [order_nums] if order_nums else []
        if not order_nums:
            order_nums = ['']

        tax_lines = invoice.get('tax_lines', [])
        if not tax_lines or not isinstance(tax_lines, list):
            tax_lines = [{
                'base_amount': invoice.get('base_amount'),
                'tax_rate': invoice.get('tax_rate'),
                'tax_amount': invoice.get('tax_amount'),
            }]

        sociedad = _get_sociedad(invoice.get('receiver_name', ''))

        # Build "Texto" field: proveedor - nº factura
        proveedor = invoice.get('company_name', '')
        num_factura = invoice.get('invoice_number', '')
        texto = f"{proveedor} - {num_factura}" if proveedor and num_factura else proveedor or num_factura or ''

        # Format invoice date to DD/MM/YYYY
        fecha_factura = _format_date_ddmmyyyy(invoice.get('date', ''))

        for on in order_nums:
            for tl in tax_lines:
                row = dict(invoice)
                row['date'] = fecha_factura
                row['order_number'] = str(on) if on else ''
                row['base_amount'] = tl.get('base_amount')
                row['tax_rate'] = tl.get('tax_rate')
                row['tax_amount'] = tl.get('tax_amount')
                row['sociedad'] = sociedad
                row['fecha_contab'] = today
                row['texto'] = texto
                row['cuenta_proveedor'] = ''  # Always empty - user fills in SAP
                row['cuenta_mayor'] = ''  # Always empty - user fills in SAP
                expanded.append(row)

    return expanded


# ─── Column definitions for each format ────────────────────────────────────

# CON PEDIDO (MIRO)
COLUMNS_CON_PEDIDO = [
    {'key': 'sociedad', 'header': 'Cod. Empresa', 'width': 12},
    {'key': 'date', 'header': 'Fecha Factura', 'width': 14},
    {'key': 'fecha_contab', 'header': 'Fecha Contab.', 'width': 14},
    {'key': 'invoice_number', 'header': 'Referencia', 'width': 16},
    {'key': 'cuenta_proveedor', 'header': 'Cuenta Proveedor', 'width': 16},
    {'key': 'company_name', 'header': 'Nombre Proveedor', 'width': 30},
    {'key': 'tax_id', 'header': 'CIF Proveedor', 'width': 14},
    {'key': 'receiver_name', 'header': 'Receptor', 'width': 30},
    {'key': 'receiver_tax_id', 'header': 'CIF Receptor', 'width': 14},
    {'key': 'base_amount', 'header': 'Base', 'width': 12, 'format': '#,##0.00'},
    {'key': 'tax_rate', 'header': '% IVA', 'width': 8, 'format': '0"%"'},
    {'key': 'tax_amount', 'header': 'IVA', 'width': 12, 'format': '#,##0.00'},
    {'key': 'total', 'header': 'Importe', 'width': 14, 'format': '#,##0.00'},
    {'key': 'texto', 'header': 'Texto', 'width': 35},
    {'key': 'order_number', 'header': 'Numero Pedido', 'width': 16},
]

# SIN PEDIDO (FI)
COLUMNS_SIN_PEDIDO = [
    {'key': 'sociedad', 'header': 'Cod. Empresa', 'width': 12},
    {'key': 'date', 'header': 'Fecha Factura', 'width': 14},
    {'key': 'fecha_contab', 'header': 'Fecha Contab.', 'width': 14},
    {'key': 'invoice_number', 'header': 'Referencia', 'width': 16},
    {'key': 'cuenta_proveedor', 'header': 'Cuenta Proveedor', 'width': 16},
    {'key': 'company_name', 'header': 'Nombre Proveedor', 'width': 30},
    {'key': 'tax_id', 'header': 'CIF Proveedor', 'width': 14},
    {'key': 'receiver_name', 'header': 'Receptor', 'width': 30},
    {'key': 'receiver_tax_id', 'header': 'CIF Receptor', 'width': 14},
    {'key': 'base_amount', 'header': 'Base', 'width': 12, 'format': '#,##0.00'},
    {'key': 'tax_rate', 'header': 'Código IVA', 'width': 10},
    {'key': 'total', 'header': 'Importe', 'width': 14, 'format': '#,##0.00'},
    {'key': 'texto', 'header': 'Texto', 'width': 35},
    {'key': 'cuenta_mayor', 'header': 'Cuenta Mayor', 'width': 14},
]


class ExcelGenerator:
    """Generate Excel files from extracted invoice data."""

    def __init__(self):
        self.header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        self.header_fill = PatternFill('solid', start_color='0D9488')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.data_font = Font(name='Calibri', size=10, color='000000')
        self.number_font = Font(name='Calibri', size=10, color='000000')
        self.account_font = Font(name='Calibri', size=10, bold=True, color='000000')
        self.border = Border(bottom=Side(style='thin', color='E5E5E5'))

    def create_workbook(self, data: List[Dict], output_path: str, con_pedido: bool = True) -> str:
        """Create Excel workbook with format based on con_pedido flag."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas con Pedido" if con_pedido else "Facturas sin Pedido"
        columns = COLUMNS_CON_PEDIDO if con_pedido else COLUMNS_SIN_PEDIDO

        # Write headers
        for col_idx, col_config in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config['width']

        ws.freeze_panes = 'A2'

        rows = _expand_invoices(data)

        current_row = 2
        for invoice in rows:
            for col_idx, col_config in enumerate(columns, 1):
                value = invoice.get(col_config['key'])
                cell = ws.cell(row=current_row, column=col_idx)

                if col_config['key'] in ('sociedad', 'cuenta_proveedor', 'cuenta_mayor'):
                    cell.value = value if value else ''
                    cell.font = self.account_font
                    cell.alignment = Alignment(horizontal='center')
                elif 'format' in col_config and value is not None:
                    cell.value = value
                    cell.number_format = col_config['format']
                    cell.alignment = Alignment(horizontal='right')
                    cell.font = self.number_font
                else:
                    cell.value = value if value else ''
                    cell.font = self.data_font

                cell.border = self.border

            current_row += 1

        # Totals
        last_row = current_row - 1
        if last_row > 1:
            total_row = last_row + 2
            col_map = {col['header']: idx for idx, col in enumerate(columns, 1)}

            # Total label
            texto_col = col_map.get('Texto', len(columns) - 1)
            ws.cell(row=total_row, column=texto_col, value="TOTAL").font = Font(bold=True, name='Calibri', size=11, color='000000')

            # Sum Base
            for header in ('Base',):
                if header in col_map:
                    c = col_map[header]
                    letter = get_column_letter(c)
                    cell = ws.cell(row=total_row, column=c)
                    cell.value = f'=SUM({letter}2:{letter}{last_row})'
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True, name='Calibri', size=11, color='000000')

            # Sum IVA (only in con_pedido)
            if 'IVA' in col_map:
                c = col_map['IVA']
                letter = get_column_letter(c)
                cell = ws.cell(row=total_row, column=c)
                cell.value = f'=SUM({letter}2:{letter}{last_row})'
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True, name='Calibri', size=11, color='000000')

            # Sum Importe
            if 'Importe' in col_map:
                c = col_map['Importe']
                letter = get_column_letter(c)
                cell = ws.cell(row=total_row, column=c)
                cell.value = f'=SUM({letter}2:{letter}{last_row})'
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True, name='Calibri', size=12, color='000000')

        wb.save(output_path)
        return output_path

    def generate_from_invoices(self, invoices: List[Dict], output_dir: str,
                                filename: str = "facturas.xlsx", con_pedido: bool = True) -> str:
        """Generate Excel file from list of extracted invoices."""
        output_path = os.path.join(output_dir, filename)
        return self.create_workbook(invoices, output_path, con_pedido=con_pedido)


class SAPCSVGenerator:
    """Generate SAP-compatible CSV from extracted invoice data."""

    def generate_csv(self, invoices: List[Dict], output_dir: str,
                     filename: str = "facturas_sap.csv", con_pedido: bool = True) -> str:
        """Generate SAP-compatible CSV. Format depends on con_pedido flag."""
        output_path = os.path.join(output_dir, filename)
        rows = _expand_invoices(invoices)
        columns = COLUMNS_CON_PEDIDO if con_pedido else COLUMNS_SIN_PEDIDO

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            writer.writerow([col['header'] for col in columns])

            for inv in rows:
                row_data = []
                for col in columns:
                    value = inv.get(col['key'], '')
                    if 'format' in col and value is not None:
                        row_data.append(self._format_decimal(value))
                    else:
                        row_data.append(value if value else '')
                writer.writerow(row_data)

        return output_path

    def _format_decimal(self, value) -> str:
        """Format number for CSV (use comma as decimal separator for SAP)."""
        if value is None:
            return ''
        try:
            return f"{float(value):.2f}".replace('.', ',')
        except (ValueError, TypeError):
            return ''
